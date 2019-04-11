import glob
import os
import re
import statistics
import sys

from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
)

from benchmark.openpyxl_helper import (
    duplicate_col,
    save_csv_as_sheet,
    RefFormula,
    split_to_append_text,
    split_to_funcs,
    build_scatter_with_mean_stdev
)

STATS_FILE_NAME = 'stats.xlsx'

INFO_MAX_COL = 6  # amount of column for info section
INFO_EXTRA_COL = 5  # column extra in info section
INFO_END_ROW = 2  # row number that info end
DATA_BEGIN_ROW = 4  # row number that data begin
DATA_COST_COL = 6  # column number of cost
DATA_MAX_COL = 6  # amount of column for data section
DATA_NUMNODE_COL = 1  # column number of number of nodes


def combine_result(folder):
    def create_main_stats():
        ############################################################
        ##              Create Scatter Chart at main              ##
        ############################################################
        ws = wb.create_sheet('main', index=0)
        start, end = 2, DATA_MAX_COL + 1
        for i in range(start, end):
            plotting_col = (i - start) * 2 + start
            xvalues = []
            yvalues = []
            stdevs = []
            titles = []
            for j, policy in enumerate(POLICIES):
                stat_ws = wb[policy]
                xvalues.append(Reference(stat_ws, min_col=1, min_row=DATA_BEGIN_ROW + 1, max_row=stat_ws.max_row))
                yvalues.append(
                    Reference(stat_ws, min_col=plotting_col, min_row=DATA_BEGIN_ROW + 1, max_row=stat_ws.max_row))
                stdevs.append(
                    Reference(stat_ws, min_col=plotting_col + 1, min_row=DATA_BEGIN_ROW + 1, max_row=stat_ws.max_row))
                titles.append(policy)
                ############################################################
                ##   Store simple stats at main screen for easy viewing   ##
                ############################################################
                ws.cell(row=2 + j, column=1, value=policy)
                for local_col_idx, col in enumerate(range(INFO_MAX_COL + 2, INFO_MAX_COL + INFO_EXTRA_COL * 2 + 2)):
                    # Set titles
                    val = RefFormula(worksheet=policy, min_col=col, min_row=1)
                    ws.cell(row=1, column=local_col_idx + 2, value=val)
                    # Set values
                    val = RefFormula(worksheet=policy, min_col=col, min_row=2)
                    ws.cell(row=2 + j, column=local_col_idx + 2, value=val)
            col_title = wb[sheets[0]].cell(row=DATA_BEGIN_ROW, column=i).value
            ############################################
            # INJECT A DIFFERENT sampled points column
            if col_title.startswith("inv.samples"):
                xvalues = []
                yvalues = []
                stdevs = []
                if col_title == "inv.samples(con)":
                    col_start_at = 14
                elif col_title == "inv.samples(obs)":
                    col_start_at = 16
                for policy in POLICIES:
                    stat_ws = wb[policy]
                    xvalues.append(Reference(stat_ws, min_col=13, min_row=DATA_BEGIN_ROW + 1, max_row=stat_ws.max_row))
                    yvalues.append(
                        Reference(stat_ws, min_col=col_start_at, min_row=DATA_BEGIN_ROW + 1, max_row=stat_ws.max_row))
                    stdevs.append(Reference(stat_ws, min_col=col_start_at + 1, min_row=DATA_BEGIN_ROW + 1,
                                            max_row=stat_ws.max_row))
            ############################################

            chart = build_scatter_with_mean_stdev(xvalues_refs=xvalues,
                                                  yvalues_refs=yvalues,
                                                  stdev_refs=stdevs,
                                                  titles=titles)
            chart.title = col_title
            # chart.style = 13
            if col_title.startswith("inv.samples"):
                chart.x_axis.title = "Number of Sampled Points"
            else:
                chart.x_axis.title = "Number of Nodes"
            chart.y_axis.title = wb[sheets[0]].cell(row=DATA_BEGIN_ROW, column=i).value
            ws.add_chart(chart, 'A' + str(4 + 15 * (i - start)))

    os.chdir(folder)
    if os.path.isfile(STATS_FILE_NAME):
        # skip as this had been processed
        return
        pass
    print(folder)
    wb = Workbook()
    ws = wb.active
    # remove default sheet
    wb.remove(ws)

    # find all the unique name of policies exists in this dir
    POLICIES = set()
    POLICIES.update(re.search('policy=(.*)_timestamp', fn).group(1)
                    for fn in glob.glob('policy=*.csv'))
    POLICIES = [p[-31:] for p in POLICIES]
    for idx, policy in enumerate(POLICIES):
        sheets, num_rows = save_csv_as_sheet(
            wb=wb,
            filename_glob="policy=*{}*.csv".format(policy),
            get_sheetname_func=lambda fn, policy=policy: "{}_{}".format(policy,
                                                                        re.search('.*timestamp=[0-9]+-([0-9]{4})[0-9]{2}.csv',
                                                                                  fn).group(1)),
            row_filter=lambda row_idx, row, z=DATA_BEGIN_ROW:
                        row if row_idx <= z
                        else (float(x) if x != 'inf' else x for x in row))

        ws = wb.create_sheet(policy, index=idx)
        create_stats_view(ws, sheets, num_rows=num_rows)
        create_stats_time_taken(wb, raw_sheets=sheets, stat_ws=ws)
        create_rescaled_sampled_pt(wb, sheets, num_rows=num_rows, stat_ws=ws)

    create_main_stats()

    # wb.save("scatter.xlsx")

    wb.save(STATS_FILE_NAME)
    # delete other csv
    for filename in glob.glob("*.csv"):
        os.remove(filename)


def create_rescaled_sampled_pt(wb, sheets, num_rows, stat_ws):
    """Create three extra columns with sampled point as the independent variables (x-axis)"""
    stats = []
    for s in sheets:
        valid_sampled_pts = [c[0].value for c in wb[s]['A5:A' + str(num_rows)] if c[0].value is not None]
        invalid_pts_con = [c[0].value for c in wb[s]['D5:D' + str(num_rows)] if c[0].value is not None]
        invalid_pts_obs = [c[0].value for c in wb[s]['E5:E' + str(num_rows)] if c[0].value is not None]
        # print(s)
        # print(invalid_pts_con)
        import numpy as np
        valid_sampled_pts = np.array(valid_sampled_pts)
        invalid_pts_con = np.array(invalid_pts_con)
        invalid_pts_obs = np.array(invalid_pts_obs)
        invalid_pts_total = invalid_pts_con + invalid_pts_obs + valid_sampled_pts
        stats.append((invalid_pts_total, invalid_pts_con, invalid_pts_obs))
    # print(stats)
    # find max number of sampled points for this run
    max_sampled_size = 0
    for sampled_pts, _, _ in stats:
        max_sampled_size = sampled_pts[-1] if sampled_pts[-1] > max_sampled_size else max_sampled_size
    step_size = int(max_sampled_size / (num_rows - 5))
    # normalize each column to same step size
    normalized_stats_con = []
    normalized_stats_obs = []
    xvalues = np.arange(start=0, stop=max_sampled_size, step=step_size)
    for s in stats:
        normalized_stats_con.append(np.interp(xvalues, s[0], s[1]))
        normalized_stats_obs.append(np.interp(xvalues, s[0], s[2]))

    # save to stat worksheet
    stats_connection = (np.mean(normalized_stats_con, axis=0), np.std(normalized_stats_con, axis=0))
    stats_obstacle = (np.mean(normalized_stats_obs, axis=0), np.std(normalized_stats_obs, axis=0))
    stat_ws['M4'] = "Sampled pts (normalized)"
    stat_ws['N4'] = "inv.samples(con)(mean)"
    stat_ws['O4'] = "inv.samples(con)(stdev)"
    for i, row in enumerate(range(DATA_BEGIN_ROW + 1, num_rows + 1)):
        stat_ws.cell(row=row, column=13, value=xvalues[i])
        stat_ws.cell(row=row, column=14, value=stats_connection[0][i])
        stat_ws.cell(row=row, column=15, value=stats_connection[1][i])
    stat_ws['P4'] = "inv.samples(obs)(mean)"
    stat_ws['Q4'] = "inv.samples(obs)(stdev)"
    for i, row in enumerate(range(DATA_BEGIN_ROW + 1, num_rows + 1)):
        stat_ws.cell(row=row, column=16, value=stats_obstacle[0][i])
        stat_ws.cell(row=row, column=17, value=stats_obstacle[1][i])


def combine_all_results(maindir):
    all_dirs = (os.path.join(maindir, x) for x in os.listdir(maindir))
    all_dirs = [x for x in all_dirs if os.path.isdir(x)]
    print('===== Combining csv to as xlxs sheet =====')
    for subdir in all_dirs:
        combine_result(subdir)
        os.chdir(maindir)  # switch back to main folder


############################################################
##                    HELPER FUNCTIONS                    ##
############################################################

def create_stats_time_taken(wb, raw_sheets, stat_ws):
    """Create stats of the number of node (and time) when the goal node is first found"""

    def find_time_taken(ws):
        solution_found = None
        terminate = None
        for r in range(DATA_BEGIN_ROW + 1, ws.max_row + 1):
            if solution_found is None and ws.cell(row=r, column=DATA_COST_COL).value != 'inf':
                # first time the policy found a solution
                # Index 1 as node count, index 2 as node cost
                solution_found = (ws.cell(row=r, column=DATA_NUMNODE_COL).value,
                                  ws.cell(row=r, column=DATA_COST_COL).value)
        # finish! lets check the final value
        if ws.cell(row=r, column=DATA_COST_COL).value != 'inf':
            terminate = (ws.cell(row=r, column=DATA_NUMNODE_COL).value,
                         ws.cell(row=r, column=DATA_COST_COL).value)
        return solution_found, terminate

    # find all the solution stats
    stats = [[], [], [], []]
    """idx 0-2 are stats when a valid solution was first found
        idx 0 : time (num of nodes)
            1 : cost
            2 : percentage of success
        idx 3-5 are same as before, but stats when it terminates"""
    for sheetname in raw_sheets:
        ws = wb[sheetname]
        solution_found, terminate = find_time_taken(ws)
        if solution_found is None:
            solution_found = [None] * 2
        else:
            stats[0].append(solution_found[0])
            stats[1].append(solution_found[1])
        if terminate is None:
            terminate = [None] * 2
        else:
            stats[2].append(terminate[0])
            stats[3].append(terminate[1])
        ws.cell(row=1, column=INFO_MAX_COL + 2, value='FirstSol node')
        ws.cell(row=2, column=INFO_MAX_COL + 2, value=solution_found[0])
        ws.cell(row=1, column=INFO_MAX_COL + 3, value='FirstSol cost')
        ws.cell(row=2, column=INFO_MAX_COL + 3, value=solution_found[1])
        ws.cell(row=1, column=INFO_MAX_COL + 4, value='TermSol node')
        ws.cell(row=2, column=INFO_MAX_COL + 4, value=terminate[0])
        ws.cell(row=1, column=INFO_MAX_COL + 5, value='TermSol cost')
        ws.cell(row=2, column=INFO_MAX_COL + 5, value=terminate[1])
    main_stats = []
    for nodesnum, cost in ((0, 1), (2, 3)):
        main_stats.append(statistics.mean(stats[nodesnum]) if len(stats[nodesnum]) > 0 else -1)
        main_stats.append(statistics.stdev(stats[nodesnum]) if len(stats[nodesnum]) > 1 else -1)
        main_stats.append(statistics.mean(stats[cost]) if len(stats[cost]) > 0 else -1)
        main_stats.append(statistics.stdev(stats[cost]) if len(stats[cost]) > 1 else -1)
        main_stats.append(str(len(stats[nodesnum]) / len(raw_sheets) * 100) + "%")
    # create stats at main page
    stat_ws['H1'] = 'FS nodes(mean)'
    stat_ws['I1'] = 'FS nodes(stdev)'
    stat_ws['J1'] = 'FS cost(mean)'
    stat_ws['K1'] = 'FS cost(stdev)'
    stat_ws['L1'] = 'FS succ (%)'
    stat_ws['M1'] = 'TS nodes(mean)'
    stat_ws['N1'] = 'TS nodes(stdev)'
    stat_ws['O1'] = 'TS cost(mean)'
    stat_ws['P1'] = 'TS cost(stdev)'
    stat_ws['Q1'] = 'TS succ (%)'

    col = INFO_MAX_COL + 2
    for i, val in enumerate(main_stats):
        stat_ws.cell(row=2, column=col, value=val)
        col += 1


def create_stats_view(main_ws, sheets, num_rows):
    ############################################################
    ##             first 4 rows just mimic others             ##
    ############################################################
    for column in range(1, INFO_MAX_COL + 1):
        for row in range(1, INFO_END_ROW + 1):
            val = RefFormula(worksheet=sheets[0],
                             min_col=column, min_row=row)

            main_ws.cell(row=row, column=column, value=val)
            # main_ws[column_letter + str(row)] =
    ############################################################
    ##    The reset of the rows should be average and stdev   ##
    ############################################################
    ## Title row
    row = 4
    # First title in column A:
    val = RefFormula(worksheet=sheets[0],
                     min_col=1, min_row=row)

    main_ws.cell(row=row, column=1, value=val)
    # others

    duplicate_col(start=2,
                  end=DATA_MAX_COL + 1,
                  ws=main_ws,
                  row=row,
                  name_generator=lambda sh=sheets[0],
                                        data=('(mean)', '(stdev)')
                  : split_to_append_text(sh, data),
                  multiples=2)
    ############################################################
    ## Values
    for row in range(DATA_BEGIN_ROW + 1, num_rows + 1):
        val = "='{sheet}'!{c}{r}".format(sheet=sheets[0],
                                         c='A', r=row)
        main_ws.cell(row=row, column=1, value=val)
        # make this lag behind as we need double the amount of original columns number
        duplicate_col(start=2,
                      end=DATA_MAX_COL + 1,
                      ws=main_ws,
                      row=row,
                      name_generator=lambda sheet_begin=sheets[0],
                                            sheet_end=sheets[-1],
                                            data=('AVERAGE', 'STDEV')
                      : split_to_funcs(sheet_begin, sheet_end, data),
                      multiples=2)


if __name__ == '__main__':
    CUR_PATH = os.path.dirname(sys.argv[0])
    dirname = os.path.abspath(os.path.join(CUR_PATH, '..', 'benchmark_output'))
    combine_all_results(dirname)
