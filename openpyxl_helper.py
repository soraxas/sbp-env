import glob
import csv

import itertools
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart import data_source

############################################################
##                    Excel Mainpulate                    ##
############################################################

def duplicate_col(start, end, ws, row, name_generator, multiples):
    """
    Given start and end of column number (end is not inclusive)
    worksheet as the target worksheet, it will duplicate columns into
    the desire multiples. e.g. if multiple is 3, 5 columns will becomes 15.
    It will obtain value of cells from generator.
    """
    # convert to 0-based system
    get_name = name_generator()
    get_name.send(None)
    for column in range(start, end):
        for i in range(multiples):
            val = get_name.send((column, row))
            dup_col = start + (column - start)*multiples + i
            ws.cell(row=row, column=dup_col, value=val)


def save_csv_as_sheet(wb, filename_glob, get_sheetname_func=None, row_filter=None):
    """
    Search through all files in directory with the filename_glob,
    if it matches, add it into workbook (wb) with a defined sheet name
    """
    if get_sheetname_func is None:
        # get it to be just the same as the sheet name
        get_sheetname_func = lambda *_ : filename
    # create a list to store the newly created sheets
    sheets = []
    if len(glob.glob(filename_glob)) < 1:
        raise ValueError("Given filename_glob is empty (None found) in current dir: {}".format(filename_glob))
    for filename in glob.glob(filename_glob):
        newsheet = get_sheetname_func(filename)
        ws = wb.create_sheet(newsheet)
        sheets.append(newsheet)
        with open(filename) as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                row_idx += 1 # conver to 1-based system
                if row_filter is not None:
                    row = row_filter(row_idx=row_idx, row=row)
                ws.append(row)
    return sheets, ws.max_row

############################################################
##                     Openpyxl Utils                     ##
############################################################

def RefFormula(**kwargs):
    return "={}".format(cReference(**kwargs))

def cReference(**kwargs):
    from openpyxl.chart.reference import DummyWorksheet
    if 'worksheet' in kwargs and isinstance(kwargs['worksheet'], str):
        # convert it to dummy worksheet
        kwargs['worksheet'] = DummyWorksheet(kwargs['worksheet'])

    # handle special case of worksheet range
    if 'worksheet_min' in kwargs:
        ws_min = kwargs.pop('worksheet_min')
        ws_max = kwargs.pop('worksheet_max')
        place_holder = 'XXXXX-XXXXX-XXXXX'
        kwargs['worksheet'] = DummyWorksheet(place_holder)
        val = "{}".format(Reference(**kwargs))
        val = val.replace(place_holder, "{}:{}".format(ws_min,ws_max))
    else:
        val = Reference(**kwargs)
    return val

############################################################
##                      Build Graphs                      ##
############################################################

def build_scatter_with_mean_stdev(xvalues_refs, yvalues_refs, stdev_refs, titles, errDir='y'):
    """Given x values, a list of y values, and the y values' corresponding
    stdev, it will return a scatter chart with stdev as error bars"""
    if len(yvalues_refs) != len(stdev_refs) != len(titles):
        raise ValueError("y-values and stdev list length must be the same")
    chart = ScatterChart()
    for xvalues, yvalues, stdev, title in zip(xvalues_refs, yvalues_refs, stdev_refs, titles):
        # convert reference to data source
        stdev_data_source = data_source.NumDataSource(numRef=data_source.NumRef(f=stdev))
        error_bars = ErrorBars(errDir=errDir, errValType='cust', plus=stdev_data_source, minus=stdev_data_source)

        series = Series(yvalues, xvalues, title=title)
        series.errBars = error_bars
        chart.series.append(series)
    return chart

############################################################
##                       GENERATOR                        ##
############################################################

def split_to_append_text(sheet, data):
    (c,r) = yield
    for i in itertools.count():
        val = RefFormula(worksheet=sheet,
                        min_col=c, min_row=r)
        val = "{}&\"{}\"".format(val, data[i%len(data)])
        (c,r) = yield val

def split_to_funcs(sheet_begin, sheet_end, data):
    (c,r) = yield
    for i in itertools.count():
        val = cReference(worksheet_min=sheet_begin,
                         worksheet_max=sheet_end,
                        min_col=c, min_row=r)
        val = "={func}({ref})".format(func=data[i%len(data)],
                                     ref=val)
        (c,r) = yield val
